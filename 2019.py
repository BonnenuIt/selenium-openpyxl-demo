from selenium import webdriver
import time
import random
from openpyxl import load_workbook

student_number_table='student_number.xlsx'
file_destination='student_score_2019_EEE.xlsx'
url_data='https://jinshuju.net/xxx'

workbook = load_workbook(student_number_table)
worksheet = workbook['Sheet1']
workbook1 = load_workbook(file_destination)
worksheet1=workbook1.active

browser = webdriver.Chrome()
for i in range(2,241):
# for i in range(262,270):
    number=worksheet.cell(row=i, column=1).value
    ID_student=worksheet.cell(row=i, column=2).value
    url = url_data
    browser.get(url)#打开浏览器预设网址
    browser.find_element_by_id('q_0_field_1').clear()
    browser.find_element_by_id('q_0_field_1').send_keys(number)
    browser.find_element_by_id('q_0_field_3').clear()
    browser.find_element_by_id('q_0_field_3').send_keys(ID_student)#password
    browser.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div[2]/form/div[2]/input').click()
    randomnumber=random.randint(0,400)/100
    time.sleep(1+randomnumber)
    # data_1=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody').text
    try:
        stdnumber=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[1]/td[2]').text
        
        stdname=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[2]/td[2]').text
        stdID=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[3]/td[2]').text
        std1=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[4]/td[2]').text
        std2=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[5]/td[2]').text
        std3=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[6]/td[2]').text
        std4=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[7]/td[2]').text
        std5=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[8]/td[2]').text
        # std6=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[9]/td[2]').text
        # std7=browser.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[2]/div[2]/div/table/tbody/tr[10]/td[2]').text
        worksheet1.cell(row = i, column = 1).value = stdnumber
        worksheet1.cell(row = i, column = 2).value = stdname
        worksheet1.cell(row = i, column = 3).value = stdID
        worksheet1.cell(row = i, column = 4).value = std1
        worksheet1.cell(row = i, column = 5).value = std2
        worksheet1.cell(row = i, column = 6).value = std3
        worksheet1.cell(row = i, column = 7).value = std4
        worksheet1.cell(row = i, column = 8).value = std5
        # worksheet1.cell(row = i, column = 9).value = std6
        # worksheet1.cell(row = i, column = 10).value = std7
    except:
        print(number,'查无此人')

workbook1.save(filename=file_destination)
print('finish')
browser.close()#关闭浏览器